"""
数据导出 API - 导出照片数据、鸟类检测数据为 Excel/CSV 格式
"""
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Optional
import csv

from app.db import database

router = APIRouter()

# 样式定义
BOLD_FONT = Font(bold=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def apply_header_style(cell):
    """应用表头样式"""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGNMENT
    cell.border = THIN_BORDER


def apply_cell_style(cell):
    """应用单元格样式"""
    cell.border = THIN_BORDER


@router.get("/photos")
async def export_photos(
    session_id: Optional[int] = Query(None, description="会话 ID，为空则导出所有照片"),
    format: str = Query("xlsx", description="导出格式：xlsx 或 csv")
):
    """
    导出照片数据

    - **session_id**: 可选，指定会话 ID 则只导出该会话的照片
    - **format**: 导出格式，支持 xlsx 或 csv
    """
    # 获取照片数据
    if session_id:
        photos = database.get_photos_by_session(session_id)
    else:
        photos = database.get_all_photos(limit=10000, offset=0)

    if not photos:
        raise HTTPException(status_code=404, detail="没有找到照片数据")

    # 定义列
    columns = [
        ("ID", "id"),
        ("会话 ID", "session_id"),
        ("文件路径", "file_path"),
        ("文件名", "file_name"),
        ("拍摄时间", "capture_time"),
        ("纬度", "gps_lat"),
        ("经度", "gps_lng"),
        ("海拔", "gps_alt"),
        ("相机型号", "camera_model"),
        ("镜头型号", "lens_model"),
        ("焦距", "focal_length"),
        ("ISO", "iso"),
        ("光圈", "aperture"),
        ("快门速度", "shutter_speed"),
        ("质量评分", "quality_score"),
        ("推荐", "is_recommended"),
        ("评分说明", "scoring_explanation"),
    ]

    if format.lower() == "csv":
        # 生成 CSV
        output = BytesIO()
        writer = csv.writer(output)
        writer.writerow([col[0] for col in columns])

        for photo in photos:
            row = []
            for _, col_name in columns:
                value = photo.get(col_name, "")
                if col_name == "is_recommended":
                    value = "是" if value else "否"
                row.append(value)
            writer.writerow(row)

        output.seek(0)
        filename = f"photos_export.csv" if not session_id else f"photos_session_{session_id}.csv"

        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    else:
        # 生成 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "照片数据"

        # 写入表头
        for col_idx, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            apply_header_style(cell)

        # 写入数据
        for row_idx, photo in enumerate(photos, 2):
            for col_idx, (_, col_name) in enumerate(columns, 1):
                value = photo.get(col_name, "")
                if col_name == "is_recommended":
                    value = "是" if value else "否"
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                apply_cell_style(cell)

        # 自动调整列宽
        for col_idx, _ in enumerate(columns, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 15

        # 输出到 BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"photos_export.xlsx" if not session_id else f"photos_session_{session_id}.xlsx"

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


@router.get("/detections")
async def export_detections(
    session_id: Optional[int] = Query(None, description="会话 ID，为空则导出所有检测"),
    format: str = Query("xlsx", description="导出格式：xlsx 或 csv")
):
    """
    导出鸟类检测数据

    - **session_id**: 可选，指定会话 ID 则只导出该会话的检测
    - **format**: 导出格式，支持 xlsx 或 csv
    """
    # 获取检测数据
    if session_id:
        detections = database.get_detections_by_session(session_id)
    else:
        # 获取所有检测（关联照片信息）
        conn = database.get_connection()
        c = conn.cursor()
        c.execute("""
            SELECT bd.*, p.file_path, p.capture_time, p.gps_lat, p.gps_lng
            FROM bird_detections bd
            JOIN photos p ON bd.photo_id = p.id
            ORDER BY bd.detection_time DESC
            LIMIT 10000
        """)
        detections = [dict(row) for row in c.fetchall()]
        conn.close()

    if not detections:
        raise HTTPException(status_code=404, detail="没有找到鸟类检测数据")

    # 定义列
    columns = [
        ("检测 ID", "id"),
        ("照片 ID", "photo_id"),
        ("物种名 (中文)", "species_cn"),
        ("物种名 (英文)", "species_en"),
        ("学名", "scientific_name"),
        ("置信度", "confidence"),
        ("描述", "description"),
        ("行为", "behavior"),
        ("文件路径", "file_path"),
        ("拍摄时间", "capture_time"),
        ("纬度", "gps_lat"),
        ("经度", "gps_lng"),
        ("检测时间", "detection_time"),
    ]

    if format.lower() == "csv":
        # 生成 CSV
        output = BytesIO()
        writer = csv.writer(output)
        writer.writerow([col[0] for col in columns])

        for detection in detections:
            row = [detection.get(col[1], "") for col in columns]
            writer.writerow(row)

        output.seek(0)
        filename = f"detections_export.csv" if not session_id else f"detections_session_{session_id}.csv"

        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    else:
        # 生成 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "鸟类检测"

        # 写入表头
        for col_idx, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            apply_header_style(cell)

        # 写入数据
        for row_idx, detection in enumerate(detections, 2):
            for col_idx, (_, col_name) in enumerate(columns, 1):
                value = detection.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                apply_cell_style(cell)

        # 自动调整列宽
        column_widths = [12, 12, 20, 20, 25, 10, 40, 30, 30, 20, 15, 15, 20]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # 输出到 BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"detections_export.xlsx" if not session_id else f"detections_session_{session_id}.xlsx"

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


@router.get("/summary")
async def export_summary(
    session_id: int = Query(..., description="会话 ID"),
    format: str = Query("xlsx", description="导出格式：xlsx 或 csv")
):
    """
    导出会话摘要

    - **session_id**: 必填，会话 ID
    - **format**: 导出格式，支持 xlsx 或 csv
    """
    # 获取会话摘要
    summary_data = database.get_session_summary(session_id)

    if not summary_data:
        raise HTTPException(status_code=404, detail="会话不存在")

    session = summary_data.get("session", {})
    stats = summary_data.get("stats", {})

    if format.lower() == "csv":
        # 生成 CSV（简化版）
        output = BytesIO()
        writer = csv.writer(output)

        # 基本信息
        writer.writerow(["=== 会话信息 ==="])
        writer.writerow(["会话 ID", session.get("id", "")])
        writer.writerow(["文件夹路径", session.get("folder_path", "")])
        writer.writerow(["扫描日期", session.get("scan_date", "")])
        writer.writerow(["地点", session.get("location_name", "")])
        writer.writerow(["备注", session.get("notes", "")])
        writer.writerow([])

        # 统计信息
        writer.writerow(["=== 统计信息 ==="])
        writer.writerow(["总照片数", stats.get("total_photos", 0)])
        writer.writerow(["推荐照片数", stats.get("recommended_count", 0)])
        writer.writerow(["有 GPS 的照片", stats.get("photos_with_gps", 0)])
        writer.writerow(["物种数量", stats.get("species_count", 0)])
        writer.writerow(["物种列表", stats.get("species_list", "")])

        output.seek(0)
        filename = f"summary_session_{session_id}.csv"

        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    else:
        # 生成 Excel
        wb = Workbook()

        # 删除默认 sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # 创建摘要 sheet
        ws_summary = wb.create_sheet("摘要")

        # 会话信息
        ws_summary.merge_cells("A1:B1")
        cell = ws_summary["A1"]
        cell.value = "=== 会话信息 ==="
        cell.font = BOLD_FONT

        summary_data_list = [
            ("会话 ID", session.get("id", "")),
            ("文件夹路径", session.get("folder_path", "")),
            ("扫描日期", session.get("scan_date", "")),
            ("地点", session.get("location_name", "")),
            ("备注", session.get("notes", "")),
        ]

        for row_idx, (label, value) in enumerate(summary_data_list, 2):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)

        # 统计信息
        ws_summary.merge_cells("A8:B8")
        cell = ws_summary["A8"]
        cell.value = "=== 统计信息 ==="
        cell.font = BOLD_FONT

        stats_data_list = [
            ("总照片数", stats.get("total_photos", 0)),
            ("推荐照片数", stats.get("recommended_count", 0)),
            ("有 GPS 的照片", stats.get("photos_with_gps", 0)),
            ("物种数量", stats.get("species_count", 0)),
            ("物种列表", stats.get("species_list", "")),
        ]

        for row_idx, (label, value) in enumerate(stats_data_list, 9):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)

        # 创建物种统计 sheet（按出现频率）
        ws_species = wb.create_sheet("物种统计")

        # 获取物种出现次数
        conn = database.get_connection()
        c = conn.cursor()
        c.execute("""
            SELECT bd.species_cn, bd.species_en, bd.scientific_name, COUNT(*) as count
            FROM bird_detections bd
            JOIN photos p ON bd.photo_id = p.id
            WHERE p.session_id = ?
            GROUP BY bd.species_cn, bd.species_en, bd.scientific_name
            ORDER BY count DESC
        """, (session_id,))
        species_rows = c.fetchall()
        conn.close()

        # 写入表头
        headers = ["物种名 (中文)", "物种名 (英文)", "学名", "出现次数"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_species.cell(row=1, column=col_idx, value=header)
            apply_header_style(cell)

        # 写入数据
        for row_idx, row in enumerate(species_rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_species.cell(row=row_idx, column=col_idx, value=value)
                apply_cell_style(cell)

        # 调整列宽
        ws_species.column_dimensions["A"].width = 20
        ws_species.column_dimensions["B"].width = 20
        ws_species.column_dimensions["C"].width = 25
        ws_species.column_dimensions["D"].width = 12

        # 输出到 BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"summary_session_{session_id}.xlsx"

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
